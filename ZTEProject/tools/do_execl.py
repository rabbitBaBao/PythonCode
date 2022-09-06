"""
@Name: do_execl.py
@Auth: 黄家健
@Date: 2022/8/29-20:15
"""
from openpyxl import load_workbook
from tools.project_path import tset_case_path, config_path
from config.read_config import ReadConfig
import re


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.workbook_object = load_workbook(filename=file_name)
        self.sheet_object = self.workbook_object[sheet_name]

    def get_excel_test_case(self, mode='all'):
        cases_list = []

        datas = list(self.sheet_object.iter_rows(values_only=True))  # 获取Excel表中的所有数据，按行显示，先是第一行的内容

        # 将Excel表中的数据拼成字典

        case_title = datas[0]  # 获取表头
        if mode == 'all':
            case_datas = datas[1:]  # 获取表数据
        else:   # 根据配置文件选择部分用例
            case_datas = []
            for i in eval(mode):
                case_datas.append(datas[i])
        for case in case_datas:
            result = dict(zip(case_title, case))
            cases_list.append(result)
        self.close_file()
        # (cases_list)
        return cases_list

    def close_file(self):
        self.workbook_object.close()


if __name__ == '__main__':
    # 测试根据配置文件选择性读取数据
    res = ReadConfig.read_config(config_path, 'Mode', 'queryPetInfo')
    cl = DoExcel(tset_case_path, 'queryPetInfo').get_excel_test_case(res)
    print(cl)

    # 测试利用正则表达式搜索替换字符串
    data = {'name': '${name1}', 'status': '${status}'}
    # post_data = DoExcel(tset_case_path, 'stringReplace').get_excel_test_case()
    # print(post_data)
    # data_replace = {}
    # for data_item in post_data:
    #     data_replace = data_item
    # print(data_replace)
    data_replace = {'name1': '影游于野', 'name2': '归鞘不归', 'status': 'available'}

    # res = re.search('\$\{(.*?)\}', data['name'])
    # print(type(res.group(0)))
    # print(res.group(0))
    # print(res.group(1))
    # data['name'] = data_replace[res.group(1)]
    # print(data)

    for key in data.keys():
        res = re.search('\$\{(.*?)\}', data[key])
        data[key] = data_replace[res.group(1)]
    print(data)
